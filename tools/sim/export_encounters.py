#!/usr/bin/env python3
"""
HearthCraft — Encounter Exporter (v2, multi-stage aware)
Reads HearthCraft_Encounter_Builder.xlsx and writes one <id>.json per encounter tab
into ./encounters/ , ready to drop into the game's assets.

Output is UNIFORM: every encounter is { ...metadata, "resupply", "stages": [ ... ] }.
- New multi-stage tabs (column-per-stage layout, marked by a "STAGES" row) export their
  ordered stage columns.
- Legacy flat tabs (one fight, fields down col C) are auto-wrapped as a single stage
  (resupply -> "none", objective derived from type, label -> name).

Usage:  python3 export_encounters.py [path_to_xlsx]
Requires: openpyxl
"""
import sys, json, os
from openpyxl import load_workbook

SKIP_TABS = {'Schema (Contract)', 'Export Help',
             'TEMPLATE — copy me', 'TEMPLATE — multi-stage'}

ENC_KEYS   = ['id', 'name', 'region', 'tier', 'recLevel', 'resupply', 'rewards', 'flavorIntro']
ENC_INT    = {'recLevel'}
STAGE_KEYS = ['stageId', 'label', 'type', 'objective', 'durationSec', 'resolve', 'drain',
              'spike', 'spikeIntervalSec', 'physMitPct', 'dread', 'shadow', 'disease',
              'cold', 'heat', 'wakefulness', 'stageFlavor']
STAGE_INT  = {'durationSec', 'resolve', 'drain', 'spike', 'spikeIntervalSec', 'physMitPct',
              'dread', 'shadow', 'disease', 'cold', 'heat', 'wakefulness'}
MAX_STAGES = 12


def _int(v):
    try:
        return int(v)
    except (TypeError, ValueError):
        return 0


def _str(v):
    return '' if v is None else str(v)


def _keyrows(ws):
    rows = {}
    for r in range(1, ws.max_row + 1):
        k = ws.cell(r, 2).value
        if k and k != 'JSON key':
            rows[k] = r
    return rows


def _is_multistage(ws, keyrows):
    for r in range(1, ws.max_row + 1):
        a = ws.cell(r, 1).value
        if isinstance(a, str) and a.startswith('STAGES'):
            return True
    return 'resupply' in keyrows and ('stageId' in keyrows or 'label' in keyrows)


def _build_stage(values):
    st = {}
    for k in STAGE_KEYS:
        st[k] = _int(values.get(k)) if k in STAGE_INT else _str(values.get(k))
    return st


def encounter_from_sheet(ws):
    keyrows = _keyrows(ws)

    if _is_multistage(ws, keyrows):
        enc = {}
        for k in ENC_KEYS:
            if k in keyrows:
                v = ws.cell(keyrows[k], 3).value
                enc[k] = _int(v) if k in ENC_INT else _str(v)
        enc.setdefault('resupply', 'none')

        stage_rows = {k: keyrows[k] for k in STAGE_KEYS if k in keyrows}
        ncols = 0
        for c in range(3, 3 + MAX_STAGES):
            if any(ws.cell(rr, c).value not in (None, '') for rr in stage_rows.values()):
                ncols = c - 2
            else:
                break
        stages = []
        for i in range(ncols):
            col = 3 + i
            vals = {k: ws.cell(rr, col).value for k, rr in stage_rows.items()}
            stages.append(_build_stage(vals))
        enc['stages'] = stages
        return enc if enc.get('id') and enc['stages'] else None

    # ---- legacy flat tab -> wrap as single stage ----
    flat = {}
    for r in range(1, ws.max_row + 1):
        k = ws.cell(r, 2).value
        if not k or k == 'JSON key':
            continue
        flat[k] = ws.cell(r, 3).value
    if not flat.get('id'):
        return None

    enc = {}
    for k in ENC_KEYS:
        if k == 'resupply':
            enc[k] = 'none'
        elif k == 'recLevel':
            enc[k] = _int(flat.get('recLevel'))
        else:
            enc[k] = _str(flat.get(k))
    typ = flat.get('type') or 'attrition'
    stage = {'stageId': 'main', 'label': flat.get('name') or flat.get('id'),
             'type': typ, 'objective': 'survive' if typ == 'survival' else 'kill'}
    for k in STAGE_KEYS:
        if k in stage:
            continue
        stage[k] = _int(flat.get(k)) if k in STAGE_INT else _str(flat.get(k))
    enc['stages'] = [stage]
    return enc


def main():
    path = sys.argv[1] if len(sys.argv) > 1 else 'HearthCraft_Encounter_Builder.xlsx'
    wb = load_workbook(path, data_only=True)
    os.makedirs('encounters', exist_ok=True)
    n = 0
    for ws in wb.worksheets:
        if ws.title in SKIP_TABS:
            continue
        enc = encounter_from_sheet(ws)
        if not enc:
            continue
        fn = os.path.join('encounters', f"{enc['id']}.json")
        with open(fn, 'w') as f:
            json.dump(enc, f, indent=2, ensure_ascii=False)
        s = len(enc['stages'])
        print(f"  wrote {fn}  ({s} stage{'s' if s != 1 else ''})")
        n += 1
    print(f"Exported {n} encounter(s) to ./encounters/")


if __name__ == '__main__':
    main()
