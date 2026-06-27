#!/usr/bin/env python3
"""
Regenerates ingredients.json and recipes.json from the master spreadsheets.

Run from the repo root:
    python3 tools/generate_data.py

Source files (edit these, then run this script):
    tools/data/hearthcraft_ingredients.xlsx  →  app/…/data/ingredients.json
    tools/data/hearthcraft_food_master.xlsx  →  app/…/data/recipes.json

Adding a new ingredient category:
    Add a column called 'category' to the 'JSON Schema' sheet in
    hearthcraft_ingredients.xlsx.  Valid values:
        grain  liquid  protein  vegetable  mushroom  herb  spice
        fat  sweetener  binder  forage  special
    The script picks it up automatically.
"""

import json
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxl not found. Run:  pip install openpyxl")
    sys.exit(1)

REPO_ROOT   = Path(__file__).parent.parent
TOOLS_DATA  = Path(__file__).parent / "data"
ASSETS_DATA = REPO_ROOT / "app/src/main/assets/data"

INGREDIENTS_XLSX = TOOLS_DATA / "hearthcraft_ingredients.xlsx"
FOOD_MASTER_XLSX = TOOLS_DATA / "hearthcraft_food_master.xlsx"
INGREDIENTS_JSON = ASSETS_DATA / "ingredients.json"
RECIPES_JSON     = ASSETS_DATA / "recipes.json"


# ── Ingredients ───────────────────────────────────────────────────────────────

def generate_ingredients() -> list:
    wb = openpyxl.load_workbook(INGREDIENTS_XLSX)
    ws = wb["JSON Schema"]

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col = {h: i + 1 for i, h in enumerate(headers) if h}

    ingredients = []
    for row in range(2, ws.max_row + 1):
        id_val = ws.cell(row, col["id"]).value
        if not id_val:
            continue
        id_val = str(id_val).strip()
        if id_val.startswith("─") or id_val.startswith("—"):
            continue

        name   = ws.cell(row, col["name"]).value or ""
        region = ws.cell(row, col["region"]).value or ""
        rarity = str(ws.cell(row, col["rarity"]).value or "c").lower()
        source = str(ws.cell(row, col["source"]).value or "forage").lower()

        # gatheringMode falls back to source if the column is blank
        gm_raw = ws.cell(row, col.get("gatheringMode", 0)).value if "gatheringMode" in col else None
        gathering_mode = str(gm_raw).lower() if gm_raw else source

        entry = {
            "id":           id_val,
            "name":         str(name).strip(),
            "region":       str(region).strip(),
            "rarity":       rarity,
            "source":       source,
            "gatheringMode": gathering_mode,
        }

        # Optional stat/tendency fields — omit if blank
        for field in ("primaryStat", "secondaryStat", "hazardTendency"):
            if field in col:
                val = ws.cell(row, col[field]).value
                if val:
                    entry[field] = str(val).lower()

        # category — only present once the column is added to the spreadsheet
        if "category" in col:
            val = ws.cell(row, col["category"]).value
            if val:
                entry["category"] = str(val).lower()

        notes = ws.cell(row, col["notes"]).value if "notes" in col else None
        if notes:
            entry["notes"] = str(notes).strip()

        ingredients.append(entry)

    return ingredients


# ── Recipes ───────────────────────────────────────────────────────────────────

def _parse_ingredients(raw) -> list:
    """'goodmans_taters x2, hearthgrain x1' → [{"id":…,"qty":…}, …]"""
    if not raw:
        return []
    result = []
    for part in str(raw).split(","):
        part = part.strip()
        if " x" in part.lower():
            # split on the last " x" to handle ingredient IDs that contain 'x'
            idx = part.lower().rfind(" x")
            ing_id = part[:idx].strip()
            qty_str = part[idx + 2:].strip()
            try:
                result.append({"id": ing_id, "qty": int(qty_str)})
            except ValueError:
                print(f"  WARNING: could not parse ingredient '{part}'")
    return result


def generate_recipes() -> list:
    wb = openpyxl.load_workbook(FOOD_MASTER_XLSX)
    ws = wb["JSON - Recipes"]

    headers = [ws.cell(1, c).value for c in range(1, ws.max_column + 1)]
    col = {h: i + 1 for i, h in enumerate(headers) if h}

    recipes = []
    for row in range(2, ws.max_row + 1):
        id_val = ws.cell(row, col["id"]).value
        if not id_val:
            continue
        id_val = str(id_val).strip()
        if id_val.startswith("─") or id_val.startswith("—"):
            continue

        recipe = {
            "id":          id_val,
            "name":        str(ws.cell(row, col["name"]).value or "").strip(),
            "band":        str(ws.cell(row, col["band"]).value or "all").lower().strip(),
            "class":       str(ws.cell(row, col["class"]).value or "food").lower().strip(),
            "method":      str(ws.cell(row, col["method"]).value or "simmer").lower().strip(),
            "tier":        int(ws.cell(row, col["tier"]).value or 1),
            "cookLevel":   int(ws.cell(row, col["cookLevel"]).value or 1),
            "description": str(ws.cell(row, col["description"]).value or "").strip(),
            "ingredients": _parse_ingredients(ws.cell(row, col["ingredients"]).value),
        }

        for field in ("primaryStat", "secondaryStat", "tertiaryStat", "hazardEffect"):
            if field in col:
                val = ws.cell(row, col[field]).value
                if val:
                    recipe[field] = str(val).lower()

        recipes.append(recipe)

    return recipes


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    print("HearthCraft data generator\n")

    print(f"Reading {INGREDIENTS_XLSX.name}…")
    ingredients = generate_ingredients()
    INGREDIENTS_JSON.write_text(json.dumps(ingredients, indent=2, ensure_ascii=False) + "\n")
    has_category = any("category" in i for i in ingredients)
    print(f"  → {len(ingredients)} ingredients → {INGREDIENTS_JSON}")
    if not has_category:
        print("  ⚠  No 'category' column found — add one before building the discovery system")

    print(f"\nReading {FOOD_MASTER_XLSX.name}…")
    recipes = generate_recipes()
    RECIPES_JSON.write_text(json.dumps(recipes, indent=2, ensure_ascii=False) + "\n")
    print(f"  → {len(recipes)} recipes → {RECIPES_JSON}")

    print("\nDone. Build the app to pick up the new data.")
    if not has_category:
        print("\nTodo: add a 'category' column to hearthcraft_ingredients.xlsx → JSON Schema sheet")
        print("  Valid values: grain  liquid  protein  vegetable  mushroom")
        print("                herb  spice  fat  sweetener  binder  forage  special")


if __name__ == "__main__":
    main()
